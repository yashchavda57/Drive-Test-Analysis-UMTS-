from django.shortcuts import render, redirect
import openpyxl
from django.conf import settings
from django.core.files.storage import FileSystemStorage
import pandas as pd
import pdfkit

from userinterface.models import DriveTestData
from userinterface.forms import DocumentForm
import folium
import os

def home(request):
    documents = DriveTestData.objects.all()
    return render(request, 'userinterface/home.html', { 'documents': documents })


def index(request):
    return render(request, 'userinterface/index.html')

def solutionview(request):
    return render(request, 'userinterface/solutions.html')

def pdfview(request):
    #pdfkit.from_file('DriveTest/userinterface/static/userinterface/report.html','report.pdf')
    return render(request, 'userinterface/report.html')

def mapview(request):
    def setcolor(val):
        if val > -70 and val < -65:
            return '#139413'
        elif val > -75 and val < -70:
            return '#16AB16'
        elif val > -80 and val < -75:
            return '#1AC91A'
        elif val > -83 and val < -80:
            return '#1DE01D'
        elif val > -83 and val < -80:
            return '#fcf803'

        elif val > -86and val< -83:
            return '#fcf803'
        elif val > -90 and val< -86:
            return '#fcbe03'
        elif val > -95 and val< -90:
            return '#fc9403'
        else:
            return '#FC350F'
    df = pd.read_excel(os.path.join('data/DataSet.xlsx'))
    locations = df[['Latitude', 'Longitude']]
    rscp = df[['Detected RSCP (dBm)(0)']]
    locationlist = locations.values.tolist()
    rscplist = rscp.values.tolist()
    map1 = folium.Map(location=[22.506774, 73.477264], zoom_start=13)

    for point in range(0, len(locationlist)):
        folium.CircleMarker(locationlist[point], color=setcolor(rscplist[point][0]),popup=(str(round(locationlist[point][0],3))+','+str(round(locationlist[point][1],3))), radius=4, fill=True, fill_opacity=1).add_to(map1)

    map1

    map1.save("userinterface/templates/userinterface/maps.html")

    #  Second Map
    rxlvl = df[['Rx Power (dBm)']]
    rxlvllist = rxlvl.values.tolist()

    def setcolor2(val):
        if val > -65:
            return 'green'
        elif val < -65 and val > -75:
            return 'yellow'
        else:
            return 'red'

    map2 = folium.Map(location=[22.506774, 73.477264], zoom_start=13)
    for point in range(0, len(locationlist)):
        folium.CircleMarker(locationlist[point], color=setcolor2(rxlvllist[point][0]),popup=(str(round(locationlist[point][0],3))+','+str(round(locationlist[point][1],3))), radius=4, fill=True, fill_opacity=1).add_to(map2)
    map2

    map2.save("userinterface/templates/userinterface/maps1.html")


    # Third Map

    ec_io = df[['Total Agg EcIo (dB)']]
    ec_io_list = ec_io.values.tolist()

    def setcolor3(val):
        if val > -7:
            return 'green'
        elif val < -7 and val > -10:
            return 'yellow'
        else:
            return 'red'

    map3 = folium.Map(location=[22.506774, 73.477264], zoom_start=15)
    for point in range(0, len(locationlist)):
        folium.CircleMarker(locationlist[point], color=setcolor3(ec_io_list[point][0]),popup=(str(round(locationlist[point][0],3))+','+str(round(locationlist[point][1],3))), radius=4, fill =True, fill_opacity = 1).add_to(map3)
    map3

    map3.save("userinterface/templates/userinterface/maps2.html")
    # change to {'my_map': map._repr_html_()}

    return render(request, 'userinterface/map.html')

def graphview(request):

        return render(request, 'userinterface/graph.html')


def tableview(request):
    if "GET" == request.method:
        return render(request, 'userinterface/table.html', {})
    else:
        excel_file = request.FILES["excel_file"]
        #excle_file = 'data/TEST.xlsx'
        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)
        print(wb.sheetnames)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["Sheet1"]
        print(worksheet)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)

        return render(request, 'userinterface/table.html', {"excel_data":excel_data})

def model_form_upload(request):
    if request.method == 'POST':
        form = DocumentForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('userinterface:home')
    else:
        form = DocumentForm()
    return render(request, 'userinterface/model_form_upload.html', {'form': form})

excel_file1 = os.path.join('data/TEST.xlsx')
excel_file2 = os.path.join('data/DataSet.xlsx')


def testview1(request):
        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file1)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["TableView"]
        print(worksheet)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)

        return render(request, 'userinterface/table.html', {"excel_data":excel_data})

def testview2(request):
        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file2)
        print(wb.sheetnames)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["Sheet1"]
        print(worksheet)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)

        return render(request, 'userinterface/table.html', {"excel_data":excel_data})


#def mapview(request)
